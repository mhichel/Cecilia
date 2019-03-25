#Este acceso directo es para poder renderizar las vistas
from django.shortcuts import render

#Aqui estamos importando los modelos para poder hacer consultas a los mismos
from apps.Entidades.models import persona, aplicacion, cargo, bodega, usuario, historico_creacion_persona, historico_eliminacion_persona, creacion_aplicacion, eliminacion_aplicacion

#Aqui importamos los forms creados para crear editar o eliminar los modelos
from apps.Entidades.forms import PersonaForm, HistoricoForm_creacion_persona, AplicacionForm, PersonaForm_addaplicacion, UsuarioForm, EditarPersonaForm, HistoricoForm_eliminacion_persona, AplicacionForm_eliminacion

#Con este metodo podemos hacer busquemas mas facilmente como se muestra en la vista de la busqueda general
from django.db.models import Q

#aqui exportamos la libreria la cual nos va a permitir exportar datos a excel
from openpyxl import Workbook
#aqui importanis varias cosas de la libreria de openpyxl para poder manipular y darle estetica a las celdas de el EXCEL
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, GradientFill

#Nos devuelve un objeto resultado, en este caso lo utilizaremos para un archivo de excel
from django.http.response import HttpResponse



#CON ESTAS IMPORTACIONES DE ESTOS COMANDOS, PODEMOS HACER QUE UN DICCIONARIO SE VUELVA UN TIPO DE DICCIONARIO DE CONSULTA, LOS CUALES SE UTILIZAN PARA ENVIARLOS A LOS FORMULARIOS Y GUARDARLOS EN LOS MODELOS
from django.http.request import QueryDict, MultiValueDict

# Create your views here.

#VISTA PRINCIPAL******************
def home(request):
    usuario = request
    return render(request, 'home.html', {'usuario':usuario})
#VISTA PRINCIPAL******************


#con este trigger hacemos un historico por cada persona creada o editada, lo llamadas en dichas vistas
def trigger_hisotrico(persona):
    aplicaciones = ""
    for apli in persona.aplicaciones.all():
        aplicaciones+=str(apli)+str(",")
    values = {
        'persona' : persona.nombre_completo,
        'aplicaciones' : aplicaciones,
        'cargo' : persona.cargo,
        'ubicacion' : persona.ubicacion,
    }

    form_hisotrico = HistoricoForm_creacion_persona(values)
    if form_hisotrico.is_valid():
        form_hisotrico.save()
    else:
        print("formulario no es valido")

#Esta es la vista de crear persona
def crear_persona(request):
    if request.method == "POST":
        print(request.POST)
        #aqui hacemos el array que va a contener las aplicaciones que sacamos de la request
        aplicaciones_POST = []
        #aqui hacemos el for a las aplicaciones que trajimos del AJAX y las metemos en el array
        for a in request.POST.get('aplicaciones'):
            aplicaciones_POST.append(a)

        #ahora vamos a hacer el diccionario que contendra todos los campos necesarios para que el formulario de creacion de una nueva persona
        dic_persona_nueva = {}
        #primero vamos a traer el nombre completo de la persona y lo agregamos al diccionario como una lista y con el nombre del campo del formulario exacto, para que funcione correctamente
        dic_persona_nueva['nombre_completo'] = [request.POST.get('nombre_completo')]
        #ahora vamos a traer las aplicaciones que se van a agregar a la persona y las metemos al formulario como una lista, traemos la variable o a lista que creamos anteriormente con el for
        dic_persona_nueva['aplicaciones'] = aplicaciones_POST
        #ahora traemos el valor del cargo asignado a la nueva persona y lo agregamos al diccionario
        dic_persona_nueva['cargo'] = [request.POST.get('cargo')]
        #ahora traemos el valor de la ubicacion asignada a la nueva persona y lo agregamos tambien al diccionario
        dic_persona_nueva['ubicacion'] = [request.POST.get('bodega')]

        #con esto finalizamos el diccionario con todos los campos requeridos para el formulario
        # ahora lo que tenemos que hacer es que el diccionario se convierta en un diccionario de consulta, para que el formulario lo pueda recibir y valir correctamente
        # para eso hacemos lo siguiente:

        #primero hacemos una variable la cual va a contener un diccionario de consulta que podra ser mutable, lo hacemos mutable para que despues podamos agregar la informacion que requerimos
        dic_persona_nueva_convertido = QueryDict('',mutable=True)
        #y ahora con este comando, hacemos la actualizacion del diccionario de consulta y la añadimos el diccionario que habiamos creado con los datos para el formulario
        #como resultado, tenemos el diccionario de consulta con la informacion requerida para el formulario, con esto ya acabamos el diccionario y seguimos con la llamada del formulario y la validacion del mismo
        dic_persona_nueva_convertido.update(MultiValueDict(dic_persona_nueva))

        #hacemos una variable que contendra el formulario y dentro del formulario le pasamos como parametro el diccionario de consulta que creamos anteriormente
        form_creacion_persona_nueva = PersonaForm(dic_persona_nueva_convertido)
        #ahora hacemos la validacion del formulario y si es valido entonces lo guardamos
        if form_creacion_persona_nueva.is_valid():
            #lo guardamos si es valido
            persona = form_creacion_persona_nueva.save()
            print("guardando la persona nueva")
            #luego con el siguiente comando, hacemos el registro en el historial de creaciones
            trigger_hisotrico(persona)

            #ahora vamos a hacer la creacion del diccionario de consulta, para enviarlo al formulario de nuevo usuario y lo guardamos despues
            #como las aplicaciones pueden ser mas de una, las que podemos agregarle a la persona, entonces tenemos que hacer varios ingresos al formulario de UsuarioForm
            # para lograr esto entonces tendremos que hacer un for, el cual iterara los arrays de las aplicacion en conjunto con el de los nombre de usuarios y numero de ticket para luego al final del for, guardar el formulario hasta que se acaben los datos
            #hacemos la lista, que contendra los numeros de tickets
            #con el metodo split, lo que hacemos es partir la cadena de texto cuando encuentre una , coma y asi sacamos los valores
            array_de_tickets = request.POST.get('array_de_tickets').split(',')
            #hacemos el array que contendra los nombre de usuario, tambien lo partimos con el metodo split
            array_de_usuarios = request.POST.get('array_de_usuarios').split(',')

            for aplicacion,ticket,nombre_usuario in zip(aplicaciones_POST, array_de_tickets, array_de_usuarios):
                #ahora aqui vamos a empezar a hacer el diccionario que va a contener los datos
                #y este diccionario despues lo vamos a convertir en un diccionario de consulta y lo enviaremos al formulario y lo guardaremos
                contenido = {}
                #empezamos agregando el usuario, ya que es el primer campo del formulario
                contenido['usuario'] = [nombre_usuario]
                #luego vamos a agregar la persona relacionada, y esto lo hacemos por medio de la variable persona, que creamos a la hora de guardar el formulario de PersonaForm
                # ya que cuando le pasamos a una varible la accion de guardar el formulario, pues en la variable quedara la informacion de la persona nueva, y entre ellas el id
                contenido['persona_relacionada'] = [str(persona.id)]
                #ahora vamos a agregar la aplicacion relacionada, y esto lo podemos hacer gracias al for que creamos anteriormente
                contenido['aplicacion_relacionada'] = [aplicacion]
                #despues agregamos el numero del ticket, el cual tambien podemos agregar gracias al for
                contenido['ticket'] = [ticket]

                #despues de tener toda la informacion requerida para el formulario, vamos a convertir el diccionario en un diccionario de consulta, para que el formulario lo pueda recibir y validar correctamente
                #primero hacemos una variable la cual contendra un diccionario de consulta, que es mutable, para que podamos agregar el diccionario que creamos con la informacion requerida para el formulario
                diccionario_de_consulta_usuarios = QueryDict('',mutable=True)

                #ahora vamos a actualizar el diccionario de consulta o mas bien vamos a agregar la informacion del diccionario que creamos hace un rato a este diccionario
                diccionario_de_consulta_usuarios.update(MultiValueDict(contenido))

                #ahora vamos a hacer la variable que contendra el formulario y como parametro para el formulario le pasamos el diccionario_de_consulta_usuarios
                form_de_usuarios_nuevos = UsuarioForm(diccionario_de_consulta_usuarios)
                #ahora validamos el formulario, y si esta bien, vamos a guardar el formulario
                if form_de_usuarios_nuevos.is_valid():
                    print("guardando usuario nuevo")
                    usuario = form_de_usuarios_nuevos.save()

                    #luego vamos a hacer el registro de las aplicaciones creadas en el historico de registro de aplicaciones nuevas
                    values_aplicacion_nueva = {
                        'ticket' : ticket,
                        'persona_relacionada' : str(persona.nombre_completo),
                        'aplicacion_relacionada' : str(persona.aplicaciones.get(id=aplicacion)),
                        'usuario' : nombre_usuario,
                    }

                    #ahora vamos a llamar el formulario y le pasamos los datos, para guardar en el historico de creacion de aplicaciones
                    form_historico_creacion_aplicacion = AplicacionForm(values_aplicacion_nueva)
                    if form_historico_creacion_aplicacion.is_valid():
                        form_historico_creacion_aplicacion.save()
                        print("guardando registro historico de aplicacion nueva")
                    else:
                        print("el formulario no es valido")

                else:
                    print("el formulario de usuarios nuevos no es valido")

                #luego todo esto se repetira el numero de veces que sea necesario, hasta acabar con los arrays, hasta que se hayan completado todos los registros segun las aplicaciones que se registraron en la persona

            #aqui lo que hacemos es declarar el estado de la peticion, si se guardo todo correctamente entonces el estado sera bien en string, el cual en la template, vamos a utilizar por medio de una condicional
            estado = "bien"


    else:
        form_creacion_persona_nueva = PersonaForm()
        #qui lo que indicamos es que si el estado de la peticion devuleve que solo vamos a enviar un formulario vacio, quiere decir que todavia no hemos enviado los datos, por lo que lo que vamos aneviar por estado es todavia en ingles (yet)
        estado = "yet"

    #aqui lo que vamos ha hacer es que si la persona, esta dentro de la variables locales, osea que si existe persona, quiere decir que el formulario se guardo correctamente, entcones lo que enviaremos ala template, sera el estado bien y la persona nueva para poder mostrar sus datos
    if "persona" in locals():
        return render(request, 'listar_personas/ajax_crear_persona.html', {'persona': persona, 'estado': estado})
        #de lo contrario pues vamos a enviar el estado de yet y el formulario vacio para que sea llenado
    else:
        return render(request, 'listar_personas/ajax_crear_persona.html', {'form': form_creacion_persona_nueva, 'estado': estado})




#esta es la vista para cuando vayamos a crear una aplicacion nueva, que nos pedira numera de ticket y demas
def crear_persona_aplicacion(request):
    #AQUI LO QUE HACEMOS ES QUE SI EL METODO DE LA REQUEST ES GET, PUES QUE HAGA LA CONSULTA A LA BBDD, BUSCANDO EL ID POR LA REQUEST.GET
    if request.method == "GET":
        persona_a_editar = persona.objects.get(id=request.GET.get('id'))

    #DE LO CONTRARIO, PUES QUE HAGA LA CONSULTA EN LA BBDD POR MEDIO DE EL PARAMETRO REQUEST.POST
    else:
        persona_a_editar = persona.objects.get(id=request.POST.get('id'))

    #SI EL METODO DE LA SOLICITUD ES POST, ENTRA EN ESTE TROZO DE CODIGO
    if request.method == "POST":
        print("entra al method post")
        #AQUI LO QUE HACEMOS ES UN ARRAY, EL CUAL DESPUES VAMOS A LLENAR CON LOS ID DE LAS APLICACIONES ADICIONADAS
        POST_aplicaciones = [];
        #AQUI CON ESTE FOR LO QUE HACEMOS ES COGER LA VARIABLE APLICACIONES QUE SACAMOS DE LA REQUEST.POST, Y LO QUE HACEMOS ES SEPARARLA Y LA GUARDAMOS EN EL ARRAY POST_aplicaciones
        for a in request.POST.get('aplicaciones'):
            POST_aplicaciones.append(a)
        print(POST_aplicaciones)

        #AQUI LO QUE HACEMOS ES SEPRAR TODOS LOS DATOS QUE OBTENELOS DE LA REQUEST.POST, Y LOS GUARDAMOS EN UN DICCIONARIO, PARA DESPUES PASARSELO A EL FORMULARIO CON LA INSTANCIA Y ASI PODER EDITAR LA PERSONA
        query = {}
        query['csrfmiddlewaretoken'] = [request.POST.get('csrfmiddlewaretoken')]
        query['id'] = [request.POST.get('id')]
        query['nombre_completo'] = [request.POST.get('nombre_completo')]
        query['aplicaciones'] = POST_aplicaciones
        query['cargo'] = [request.POST.get('cargo')]
        query['ubicacion'] = [request.POST.get('ubicacion')]

        #AQUI LO UQE HACEMOS ES HACER UN DICCIONADO DE CONSULTA MUTABLE POR MEDIO DE ESTAS KEYWORDS
        qdict = QueryDict('',mutable=True)
        #Y AQUI LO QUE HACEMOS ES ACTUALIZAR EL DICCIONARIO CON EL DICCIONARIO QUERY QUE HABIAMOS HECHO ANTERIORMENTE
        # Y ASI POR MEDIO DE ESTOS COMANDOS PODEMOS HACER UN DICCIONARIO DE CONSULTA, COMO EL QUE SE CREA EN LA REQUEST.POST
        qdict.update(MultiValueDict(query))

        #AQUI HACEMOS UNA VARIABLE, LA CUAL LE PASAMOS EL FORMULARIO DE EDICION DE LA PERSONA, Y COMO PARAMETROS DE ACTUALIZACION LE PASAMOS EL DICCIONARIO
        #QUE HICIMOS ANTERIORMENTE, EL DICCIONARIO DE CONSULTA LLAMADO QDICT, Y COMO SEGUNDO PARAMETRO, PASAMOS LA INSTANCIA DE LA PERSONA A EDITAR, EL CUAL OBTUBIMOS POR MEDIO DE LA PRIMERA LINEA O SEGUNDA DE LA FUNCION DEPENDE DE EL METODO DE SOLICITUD
        form_editar_persona = PersonaForm_addaplicacion(qdict, instance=persona_a_editar)

        #AQUI VERIFICAMOS SI EL FORMULARIO ES VALIDO, ENTONCES QUE LO GUARDE
        if form_editar_persona.is_valid():
            persona_1 = form_editar_persona.save()
            print("guardando edicion de persona")

            #AHORA AQUI VAMOS A GUARDAR EL FORMULARIO DE LA NUEVA APLICACION ADICIONADA
            #PRIMERO HAREMOS LA QUERY, LAS CUALES GUARDAREMOS ES UN DICCIONARIO QUE DESPUES LE PASAREMOS COMO PARAMETRO PARA QUE GUARDE LOS CAMBIOS
            data_aplicacionForm = {}
            data_aplicacionForm['ticket'] = [request.POST.get('ticket_creacion')]
            data_aplicacionForm['persona_relacionada'] = [persona_1.nombre_completo]
            data_aplicacionForm['aplicacion_relacionada'] = [persona_1.aplicaciones.get(id=request.POST.get('aplicacion_a_adicionar'))]
            data_aplicacionForm['usuario'] = [request.POST.get('usuario')]

            #LUEGO EL DICCIONARIO HECHO, LO CONVERTIMOS EN UN DICCIONARIO DE CONSULTA, PARA PODER PASARSELO AL FORMULARIO, Y PODER GUARDARLO
            data = QueryDict('',mutable=True)
            data.update(MultiValueDict(data_aplicacionForm))

            #LUEGO LE PASAMOS LOS PAREMTROS OBTENIDOS DE LA REQUEST.POST Y SE LOS PONEMOS AL FORMULARIO
            # Y SI EL FORMULARIO ES VALIDO, QUE LO GUARDE
            print(data)
            form_tabla_de_nueva_aplicacion = AplicacionForm(data)
            if form_tabla_de_nueva_aplicacion.is_valid():
                form_tabla_de_nueva_aplicacion.save()
                print("guardando registro de aplicacion nueva")
            else:
                print("el formulario no es valido")

            #AHORA AQUI VAMOS A GUARDAR LA INFORMACION DEL USUARIO EN OTRO MODELO QUE ES EL QUE GUARDARA LOS DATOS ACTUALES DE LA PERSONA Y SU RELACION DE USUARIO CON APLICACION
            #AQUI CREAMOS EL DICCIONARIO EN EL CUAL VAMOS A GARDAR TODOS LOS DATOS REQUERIDOS PARA EL FORMULARIO UsuarioForm
            data_usuarioform = {}
            data_usuarioform['usuario'] = [request.POST.get('usuario')]
            data_usuarioform['persona_relacionada'] = [request.POST.get('id')]
            data_usuarioform['aplicacion_relacionada'] = [request.POST.get('aplicacion_a_adicionar')]
            data_usuarioform['ticket'] = [request.POST.get('ticket_creacion')]

            #AHORA VAMOS A HACER QUE EL DICCIONARIO SE VUELVA UN DICCIONARIO DE CONSULTA, PARA QUE LO PODAMOS ENVIAR AL FORM
            data_2 = QueryDict('',mutable=True)
            data_2.update(MultiValueDict(data_usuarioform))

            #Y AQUI SE LO ENVIAMOS AL FORMULARIO DE UsuarioForm
            print(data_2)
            form_usuarioform = UsuarioForm(data_2)
            if form_usuarioform.is_valid():
                form_usuarioform.save()
                print("guardando informacion de usuario actual")
            else:
                print("el formulario 2 no es valido")


            # Y DESPUES QUE ESTA VARIABLE TENGA EL STRING DE BIEN, EL CUAL VAMOS A CONDICIONAR EN LA TEMPLATE
            formulario_de_crear_persona_aplicacion = "bien"
        else:
            print("formulario no es valido")
            formulario_de_crear_persona_aplicacion = "mal"
    else:
        #DE LO CONTRARIO QUE TRAIGA EL FORMULARIO CON LA INSTANCIA DE LA PERSONA A EDITAR, O SEA QUE LLENE EL FORMULARIO CON LA INFORMACION DE LA PERSONA A ADICIONAR
        formulario_de_crear_persona_aplicacion = PersonaForm_addaplicacion(instance=persona_a_editar);
    #AQUI HACEMOS UNA CONDICIONAL QUE SI LA PERSONA_1 ESTA CREADA EN LAS VARIABLE LOCALES (DE LA FUNCION), QUIERE DECIR QUE YA SE GUARDO EN LA BBDD Y PODREMOS MOSTRAR LOS CAMBIOS EN LA TEMPLATE
    if "persona_1" in locals():
        return render(request, 'listar_personas/ajax_crear_persona_aplicacion.html', {'form':formulario_de_crear_persona_aplicacion, 'persona': persona_1,})
    #DE LO CONTRARIO PUES QUE SOLAMENTE SE LLEVE AL TEMPLATE LA VARIABLE FORM, LA CUAL TIENE UN STRING DE MAL O BIEN
    else:
        return render(request, 'listar_personas/ajax_crear_persona_aplicacion.html', {'form':formulario_de_crear_persona_aplicacion})
#esta es la vista para cuando vayamos a crear una aplicacion nueva, que nos pedira numera de ticket y demas


#Este es la vista de listar el general de las personas
#OBTENEMOS TODOS LOS DATOS DE TODOS LOS MODELOS PARA ASI PODERLOS PONER TODOS EN LOS SELECTS Y EN LA TABLA
def listar_personas(request):
    lista = persona.objects.all()
    aplicaciones = aplicacion.objects.all()
    cargos = cargo.objects.all()
    bodegas = bodega.objects.all()


    return render(request, 'listar_personas/listar_personas.html', {'lista': lista, 'aplicaciones': aplicaciones,'cargos': cargos, 'bodegas':bodegas})

#Esta es la vista que utiliza filtros y ordenaciones
def busquedas_con_filtros(request):
    if request.method == "GET":
        #CON ESTO SE HACE LA BUSQUEDA GENERAL
        #SI EN LA SOLICITUD ESTA UNA VARIABLE LLAMADA "query_gral", ENTONCES QUE VALIDE LO SIGUIENTE
        if request.GET.get('query_gral'):
            #HACE LA VARIABLE, CON EL VALOR QUE LE PASAMOS AL INPUT EN LA TEMPLATE, DESDE ALLA, LO OBTENEMOS POR MEDIO DE AJAX
            query = request.GET.get('query_gral')
            #HACE UN CONJUNTO DE BUSQEUDAS EN TODOS LOS MODELOS, SEGUN EL VALOR QUE LE PASEMOS PARA BUSCAR
            qset = (
                Q(nombre_completo__icontains=query) |
                Q(aplicaciones__nombre_aplicacion__icontains=query) |
                Q(cargo__nombre_cargo__icontains=query) |
                Q(ubicacion__ubicacion__icontains=query) |
                Q(usuario__usuario__icontains=query) |
                Q(usuario__ticket__icontains=query)
                )
            #LUEGO AQUI HACE LA BUSQUEDA EN EL MODELO PERSONA Y LUEGO LO ENVIAMOS CON UN RENDER
            lista = persona.objects.filter(qset).distinct()
            print("aqui entra aplantilla de busqeuda con filtros")
            #aqui primero vamos a evluar si en la request viene una variable de esportacion a excel, pues se creara y se enviara de lo contrario pues que renderize en el html la informacion requerida en la busqueda
            if request.GET.get('exportacion_a_excel'):
                print("entra a exportacion a excel")
                #entonces aqui vamos a comenzar a crear el libro con los siguiente comandos
                #instanciamos el libro y lo llamamos libro_busqueda_general_personas
                libro_busqueda_general_personas = Workbook()
                #luego nos toca definir una hoja como activa para comenzar a editarla, entonces por defecto vamos a activar la hoja 1 por defecto
                hoja_activa_busqueda_general_personas = libro_busqueda_general_personas.active

                hoja_activa_busqueda_general_personas.page_setup.orientation = hoja_activa_busqueda_general_personas.ORIENTATION_LANDSCAPE
                hoja_activa_busqueda_general_personas.page_setup.paperSize = hoja_activa_busqueda_general_personas.PAPERSIZE_TABLOID
                hoja_activa_busqueda_general_personas.page_setup.fitToHeight = 0
                hoja_activa_busqueda_general_personas.page_setup.fitToWidth = 1



                #luego con la hoja activa vamos a trabajar
                #ahora vamos a juntar algunas celdas para que se vea que el titulo abarca toda la tabla
                hoja_activa_busqueda_general_personas.merge_cells('A1:F1')
                titulo = hoja_activa_busqueda_general_personas['A1']
                #vamos a poner el nombre de la tabla con el siguiente parametro
                titulo.value = "Gestora de entidades Cecilia"


                #ahora vamos a aplicar los estilos a la celda
                #titulo.font = Font(fuente_titulo)
                thin = Side(border_style="thin", color="000000")
                double = Side(border_style="double", color="ff0000")


                titulo.fill = PatternFill("solid", fgColor="000000")
                titulo.font  = Font(b=True, color="ffffff")
                titulo.alignment = Alignment(horizontal="center", vertical="center")




                #ahora vamos a crear los campos de la tabla, o sea los titulos de cada columna
                hoja_activa_busqueda_general_personas['A2'] = "Ubicacion"
                #luego aqui guardamos en una variable la celda para despues mas adelante con un metodo ponerle un estilo
                cell_ubicacion = hoja_activa_busqueda_general_personas['A2']

                hoja_activa_busqueda_general_personas['B2'] = "Persona"
                #luego aqui guardamos en una variable la celda para despues mas adelante con un metodo ponerle un estilo
                cell_persona = hoja_activa_busqueda_general_personas['B2']

                hoja_activa_busqueda_general_personas['C2'] = "Cargo"
                #luego aqui guardamos en una variable la celda para despues mas adelante con un metodo ponerle un estilo
                cell_cargo = hoja_activa_busqueda_general_personas['C2']

                hoja_activa_busqueda_general_personas['D2'] = "Aplicaciones"
                #luego aqui guardamos en una variable la celda para despues mas adelante con un metodo ponerle un estilo
                cell_aplicacion = hoja_activa_busqueda_general_personas['D2']

                hoja_activa_busqueda_general_personas['E2'] = "Usuario"
                #luego aqui guardamos en una variable la celda para despues mas adelante con un metodo ponerle un estilo
                cell_usuario = hoja_activa_busqueda_general_personas['E2']

                hoja_activa_busqueda_general_personas['F2'] = "Ticket creacion"
                #luego aqui guardamos en una variable la celda para despues mas adelante con un metodo ponerle un estilo
                cell_ticket = hoja_activa_busqueda_general_personas['F2']

                #aqui vamos a ponerles estilos a los titulos de las columnas para que se vea mejor
                cell_ubicacion.fill = PatternFill("solid", fgColor="000000")
                cell_ubicacion.font  = Font(b=True, color="ffffff")
                cell_persona.fill = PatternFill("solid", fgColor="000000")
                cell_persona.font  = Font(b=True, color="ffffff")
                cell_cargo.fill = PatternFill("solid", fgColor="000000")
                cell_cargo.font  = Font(b=True, color="ffffff")
                cell_aplicacion.fill = PatternFill("solid", fgColor="000000")
                cell_aplicacion.font  = Font(b=True, color="ffffff")
                cell_usuario.fill = PatternFill("solid", fgColor="000000")
                cell_usuario.font  = Font(b=True, color="ffffff")
                cell_ticket.fill = PatternFill("solid", fgColor="000000")
                cell_ticket.font  = Font(b=True, color="ffffff")

                #Ahora vamos a crear un contador para recorrer todos los registros de la variable que tiene los registros de la tabla de historicos de personas creadas
                contador_busqueda_general_personas=3

                #vamos a hacer una variable que se iterara cada vez que pase a otra fila, esto con el fin de pintarla de otro color
                color = 0
                #ahora vamos a empezar a insertar todos los registros en la hoja segun el contador, el contador va a ser la fila actual y se incrementara cada vez que se registre una nueva persona
                for registro in lista:
                    #aqui unimos las celdas segun el numero de aplicaciones que tenga la persona, esto para que sea vea mas estetico en el excel y al final le descontamos -1 por que como en programacion comienza a contar desde cero entonces sera una apicacion mas
                    hoja_activa_busqueda_general_personas.merge_cells('A'+str(contador_busqueda_general_personas)+':'+'A'+str(contador_busqueda_general_personas+registro.aplicaciones.all().count()-1))
                    hoja_activa_busqueda_general_personas.merge_cells('B'+str(contador_busqueda_general_personas)+':'+'B'+str(contador_busqueda_general_personas+registro.aplicaciones.all().count()-1))
                    hoja_activa_busqueda_general_personas.merge_cells('C'+str(contador_busqueda_general_personas)+':'+'C'+str(contador_busqueda_general_personas+registro.aplicaciones.all().count()-1))

                    #luego de unir las celdas, vamos a empezar a insertar las datos de la persona aqui estamos insertando la ubicacion
                    hoja_activa_busqueda_general_personas.cell(row=contador_busqueda_general_personas,column=1).value = str(registro.ubicacion)
                    #aqui vamos a guardar en una variable la celda para poder pintarla con el metodo fill mas delante
                    valor1 = hoja_activa_busqueda_general_personas.cell(row=contador_busqueda_general_personas,column=1)

                    #luego de unir las celdas, vamos a empezar a insertar las datos de la persona aqui estamos insertando el nombre completo
                    hoja_activa_busqueda_general_personas.cell(row=contador_busqueda_general_personas,column=2).value = str(registro.nombre_completo)
                    #aqui vamos a guardar en una variable la celda para poder pintarla con el metodo fill mas delante
                    valor2 = hoja_activa_busqueda_general_personas.cell(row=contador_busqueda_general_personas,column=2)

                    #luego de unir las celdas, vamos a empezar a insertar las datos de la persona aqui estamos insertando el cargo
                    hoja_activa_busqueda_general_personas.cell(row=contador_busqueda_general_personas,column=3).value = str(registro.cargo)
                    #aqui vamos a guardar en una variable la celda para poder pintarla con el metodo fill mas delante
                    valor3 = hoja_activa_busqueda_general_personas.cell(row=contador_busqueda_general_personas,column=3)




                    #luego aqui hacemos un contador, el cual nos servira para poder contar las aplicaciones que tiene la persona y asi saltarse de fila en fila para poder insertar cada aplicacion
                    contador_aplicaciones=0
                    #aqui hacemos un for iterando dentro del registro de la persona, vamos a la relacion de la tabla de usuarios, el cual tiene la informacion de las aplicaciones relacionadas como el usuario y el ticket
                    for info_aplicacion in registro.usuario_set.all():
                        #y aqui entramos a iterar las aplicaciones, ya que es un campo many to many
                        for aplicacion in info_aplicacion.aplicacion_relacionada.all():
                            #por ultimo comenzamos a insertar los datos de cada aplicacion que tiene la persona aqui insertamos el nombre de la aplicacion
                            hoja_activa_busqueda_general_personas.cell(row=contador_busqueda_general_personas+contador_aplicaciones,column=4).value = str(aplicacion)
                            #aqui vamos a guardar en una variable la celda para poder pintarla con el metodo fill mas delante
                            valor_aplicacion1 = hoja_activa_busqueda_general_personas.cell(row=contador_busqueda_general_personas+contador_aplicaciones,column=4)

                            #aqui insertamos el usuario relacionado con esta aplicacion
                            hoja_activa_busqueda_general_personas.cell(row=contador_busqueda_general_personas+contador_aplicaciones,column=5).value = str(info_aplicacion.usuario)
                            #aqui vamos a guardar en una variable la celda para poder pintarla con el metodo fill mas delante
                            valor_aplicacion2 = hoja_activa_busqueda_general_personas.cell(row=contador_busqueda_general_personas+contador_aplicaciones,column=5)

                            #aqui insertamos el ticket de creacion del usuario de la aplicacion
                            hoja_activa_busqueda_general_personas.cell(row=contador_busqueda_general_personas+contador_aplicaciones,column=6).value = str(info_aplicacion.ticket)
                            #aqui vamos a guardar en una variable la celda para poder pintarla con el metodo fill mas delante
                            valor_aplicacion3 = hoja_activa_busqueda_general_personas.cell(row=contador_busqueda_general_personas+contador_aplicaciones,column=6)

                            #luego aqui evaluamos si la variable color es 0 entonces que lo pinte de un color diferente que si la variable color fuera 1
                            #y pues de esta manera se pintada cada fila de un color diferente
                            if color == 0:
                                valor_aplicacion1.fill = PatternFill("solid", fgColor="bdbdbd")
                                valor_aplicacion2.fill = PatternFill("solid", fgColor="bdbdbd")
                                valor_aplicacion3.fill = PatternFill("solid", fgColor="bdbdbd")
                            else:
                                valor_aplicacion1.fill = PatternFill("solid", fgColor="a3a3a3")
                                valor_aplicacion2.fill = PatternFill("solid", fgColor="a3a3a3")
                                valor_aplicacion3.fill = PatternFill("solid", fgColor="a3a3a3")


                            #y al contador le sumamos uno, el cual sera la fila, o sea que sera una fila mas abajo
                            contador_aplicaciones=contador_aplicaciones + 1
                    #al final le sumamos al contadorgeneral el del principio, el numero de aplicaciones de la persona, para que en el momento de que se registre la nueva aplicacion, pues se registr abajo de la insertada anteriormente y no haya ni espacios ni problemas de ubicacion
                    contador_busqueda_general_personas = contador_busqueda_general_personas + registro.aplicaciones.all().count()


                    #luego aqui evaluamos si la variable color es 0 entonces que lo pinte de un color diferente que si la variable color fuera 1
                    #y pues de esta manera se pintada cada fila de un color diferente
                    if color == 0:
                        valor1.fill = PatternFill("solid", fgColor="bdbdbd")
                        valor2.fill = PatternFill("solid", fgColor="bdbdbd")
                        valor3.fill = PatternFill("solid", fgColor="bdbdbd")
                        color = 1
                    else:
                        valor1.fill = PatternFill("solid", fgColor="a3a3a3")
                        valor2.fill = PatternFill("solid", fgColor="a3a3a3")
                        valor3.fill = PatternFill("solid", fgColor="a3a3a3")
                        color = 0


                #despues de registrar los datos en la tabla, vamos a ponerle nombre al archivo
                nombre_archivo_busqueda_general_personas = "Reporte_de_personas.xlsx"

                #despues vamos a definir con los siguientes comandos, el metodo de respuesta que se enviara
                #aqui le indicamos que la respuesta va a ser un archivo de microsoft excel
                response = HttpResponse(content_type="application/ms-excel")
                #aqui especificamos el nombre del archivo
                contenido = "attachment; filename={0}".format(nombre_archivo_busqueda_general_personas)
                #aqui le declaramos a la respuesta el nombre del archivo
                response["Content-Disposition"] = contenido
                #aqui guardamos el archivo en la respuesta
                libro_busqueda_general_personas.save(response)
                # y por utlimo devolvemos la respuesta que sera el archivo
                return response
            else:
                return render(request, 'listar_personas/ajax_busquedas_con_filtros.html', {'lista': lista,'query': query})


        #CON ESTO SE HACE LA BUSQUEDA ANIDADA
        lista = persona.objects.all()
        #AQUI RELACIONAMOS LA UBICACION Y SI NO ESTA, EL OBJETO LISTA LO DEJA IGUAL SIN NINGUN FILTRO POR PARTE DE LA UBICACION
        if request.GET.get('ubicacion'):
            ubicacion = request.GET.get('ubicacion')
            if ubicacion == "all":
                lista = lista
            else:
                lista = lista.filter(ubicacion=ubicacion)
        #AQUI RELACIONAMOS EL CARGO Y SI NO ESTA, EL OBJETO LISTA LO DEJA IGUAL SIN NINGUN FILTRO POR PARTE DE CARGOS
        if request.GET.get('cargo'):
            cargo = request.GET.get('cargo')
            if cargo == "all":
                lista = lista
            else:
                lista = lista.filter(cargo=cargo)
        #AQUI RELACIONAMOS EL  NOMBRE COMPLETO Y SI NO ESTA, EL OBJETO LISTA LO DEJA IGUAL SIN NINGUN FILTRO POR PARTE DEL NOMBRE COMPLETO
        if request.GET.get('nombre_completo'):
            nombre_completo = request.GET.get('nombre_completo')
            if nombre_completo == "all":
                lista = lista
            else:
                lista = lista.filter(nombre_completo__icontains=nombre_completo)
        #AQUI RELACIONAMOS LA APLICACION Y SI NO ESTA, EL OBJETO LISTA LO DEJA IGUAL SIN NINGUN FILTRO POR PARTE DE LAS APLICACIONES
        if request.GET.get('aplicacion'):
            aplicacion = request.GET.get('aplicacion')
            if aplicacion == "all":
                lista = lista
            else:
                lista = lista.filter(aplicaciones=aplicacion)

        #AQUI HACEMOS ESTAS CONDICIONALES PARA QUE CUANDO QUITEMOS LOS FILTROS, SOLAMENTE SE QUITEN LOS FILTROS DE ORDENADO MAS NO LOS DE LOS SELECTS Y DEMAS
        if request.GET.get('direccion_filtro') == "desendente":
            lista_1 = lista.order_by('-' + request.GET.get('ordenado'))
        elif request.GET.get('direccion_filtro') == "acendente":
            lista_1 = lista.order_by(request.GET.get('ordenado'))
        else:
            lista_1 = lista
        #AQUI HACEMOS ESTAS CONDICIONALES PARA QUE CUANDO QUITEMOS LOS FILTROS, SOLAMENTE SE QUITEN LOS FILTROS DE ORDENADO MAS NO LOS DE LOS SELECTS Y DEMAS

        #AQUI EVALUAMOS LA PARTE DE SI QUITAMOS EL FILTRO O NO
        if request.GET.get('quitar_filtro') == "si":
            lista = lista
        else:
            lista = lista_1


        if request.GET.get('exportacion_a_excel'):
            print("entra a exportacion a excel")
            #entonces aqui vamos a comenzar a crear el libro con los siguiente comandos
            #instanciamos el libro y lo llamamos libro_busqueda_general_personas
            libro_busqueda_general_personas = Workbook()
            #luego nos toca definir una hoja como activa para comenzar a editarla, entonces por defecto vamos a activar la hoja 1 por defecto
            hoja_activa_busqueda_general_personas = libro_busqueda_general_personas.active

            hoja_activa_busqueda_general_personas.page_setup.orientation = hoja_activa_busqueda_general_personas.ORIENTATION_LANDSCAPE
            hoja_activa_busqueda_general_personas.page_setup.paperSize = hoja_activa_busqueda_general_personas.PAPERSIZE_TABLOID
            hoja_activa_busqueda_general_personas.page_setup.fitToHeight = 0
            hoja_activa_busqueda_general_personas.page_setup.fitToWidth = 1



            #luego con la hoja activa vamos a trabajar
            #ahora vamos a juntar algunas celdas para que se vea que el titulo abarca toda la tabla
            hoja_activa_busqueda_general_personas.merge_cells('A1:F1')
            titulo = hoja_activa_busqueda_general_personas['A1']
            #vamos a poner el nombre de la tabla con el siguiente parametro
            titulo.value = "Gestora de entidades Cecilia"


            #ahora vamos a aplicar los estilos a la celda
            #titulo.font = Font(fuente_titulo)
            thin = Side(border_style="thin", color="000000")
            double = Side(border_style="double", color="ff0000")


            titulo.fill = PatternFill("solid", fgColor="000000")
            titulo.font  = Font(b=True, color="ffffff")
            titulo.alignment = Alignment(horizontal="center", vertical="center")




            #ahora vamos a crear los campos de la tabla, o sea los titulos de cada columna
            hoja_activa_busqueda_general_personas['A2'] = "Ubicacion"
            #luego aqui guardamos en una variable la celda para despues mas adelante con un metodo ponerle un estilo
            cell_ubicacion = hoja_activa_busqueda_general_personas['A2']

            hoja_activa_busqueda_general_personas['B2'] = "Persona"
            #luego aqui guardamos en una variable la celda para despues mas adelante con un metodo ponerle un estilo
            cell_persona = hoja_activa_busqueda_general_personas['B2']

            hoja_activa_busqueda_general_personas['C2'] = "Cargo"
            #luego aqui guardamos en una variable la celda para despues mas adelante con un metodo ponerle un estilo
            cell_cargo = hoja_activa_busqueda_general_personas['C2']

            hoja_activa_busqueda_general_personas['D2'] = "Aplicaciones"
            #luego aqui guardamos en una variable la celda para despues mas adelante con un metodo ponerle un estilo
            cell_aplicacion = hoja_activa_busqueda_general_personas['D2']

            hoja_activa_busqueda_general_personas['E2'] = "Usuario"
            #luego aqui guardamos en una variable la celda para despues mas adelante con un metodo ponerle un estilo
            cell_usuario = hoja_activa_busqueda_general_personas['E2']

            hoja_activa_busqueda_general_personas['F2'] = "Ticket creacion"
            #luego aqui guardamos en una variable la celda para despues mas adelante con un metodo ponerle un estilo
            cell_ticket = hoja_activa_busqueda_general_personas['F2']

            #aqui vamos a ponerles estilos a los titulos de las columnas para que se vea mejor
            cell_ubicacion.fill = PatternFill("solid", fgColor="000000")
            cell_ubicacion.font  = Font(b=True, color="ffffff")
            cell_persona.fill = PatternFill("solid", fgColor="000000")
            cell_persona.font  = Font(b=True, color="ffffff")
            cell_cargo.fill = PatternFill("solid", fgColor="000000")
            cell_cargo.font  = Font(b=True, color="ffffff")
            cell_aplicacion.fill = PatternFill("solid", fgColor="000000")
            cell_aplicacion.font  = Font(b=True, color="ffffff")
            cell_usuario.fill = PatternFill("solid", fgColor="000000")
            cell_usuario.font  = Font(b=True, color="ffffff")
            cell_ticket.fill = PatternFill("solid", fgColor="000000")
            cell_ticket.font  = Font(b=True, color="ffffff")

            #Ahora vamos a crear un contador para recorrer todos los registros de la variable que tiene los registros de la tabla de historicos de personas creadas
            contador_busqueda_general_personas=3

            #vamos a hacer una variable que se iterara cada vez que pase a otra fila, esto con el fin de pintarla de otro color
            color = 0
            #ahora vamos a empezar a insertar todos los registros en la hoja segun el contador, el contador va a ser la fila actual y se incrementara cada vez que se registre una nueva persona
            for registro in lista:
                #aqui unimos las celdas segun el numero de aplicaciones que tenga la persona, esto para que sea vea mas estetico en el excel y al final le descontamos -1 por que como en programacion comienza a contar desde cero entonces sera una apicacion mas
                hoja_activa_busqueda_general_personas.merge_cells('A'+str(contador_busqueda_general_personas)+':'+'A'+str(contador_busqueda_general_personas+registro.aplicaciones.all().count()-1))
                hoja_activa_busqueda_general_personas.merge_cells('B'+str(contador_busqueda_general_personas)+':'+'B'+str(contador_busqueda_general_personas+registro.aplicaciones.all().count()-1))
                hoja_activa_busqueda_general_personas.merge_cells('C'+str(contador_busqueda_general_personas)+':'+'C'+str(contador_busqueda_general_personas+registro.aplicaciones.all().count()-1))

                #luego de unir las celdas, vamos a empezar a insertar las datos de la persona aqui estamos insertando la ubicacion
                hoja_activa_busqueda_general_personas.cell(row=contador_busqueda_general_personas,column=1).value = str(registro.ubicacion)
                #aqui vamos a guardar en una variable la celda para poder pintarla con el metodo fill mas delante
                valor1 = hoja_activa_busqueda_general_personas.cell(row=contador_busqueda_general_personas,column=1)

                #luego de unir las celdas, vamos a empezar a insertar las datos de la persona aqui estamos insertando el nombre completo
                hoja_activa_busqueda_general_personas.cell(row=contador_busqueda_general_personas,column=2).value = str(registro.nombre_completo)
                #aqui vamos a guardar en una variable la celda para poder pintarla con el metodo fill mas delante
                valor2 = hoja_activa_busqueda_general_personas.cell(row=contador_busqueda_general_personas,column=2)

                #luego de unir las celdas, vamos a empezar a insertar las datos de la persona aqui estamos insertando el cargo
                hoja_activa_busqueda_general_personas.cell(row=contador_busqueda_general_personas,column=3).value = str(registro.cargo)
                #aqui vamos a guardar en una variable la celda para poder pintarla con el metodo fill mas delante
                valor3 = hoja_activa_busqueda_general_personas.cell(row=contador_busqueda_general_personas,column=3)




                #luego aqui hacemos un contador, el cual nos servira para poder contar las aplicaciones que tiene la persona y asi saltarse de fila en fila para poder insertar cada aplicacion
                contador_aplicaciones=0
                #aqui hacemos un for iterando dentro del registro de la persona, vamos a la relacion de la tabla de usuarios, el cual tiene la informacion de las aplicaciones relacionadas como el usuario y el ticket
                for info_aplicacion in registro.usuario_set.all():
                    #y aqui entramos a iterar las aplicaciones, ya que es un campo many to many
                    for aplicacion in info_aplicacion.aplicacion_relacionada.all():
                        #por ultimo comenzamos a insertar los datos de cada aplicacion que tiene la persona aqui insertamos el nombre de la aplicacion
                        hoja_activa_busqueda_general_personas.cell(row=contador_busqueda_general_personas+contador_aplicaciones,column=4).value = str(aplicacion)
                        #aqui vamos a guardar en una variable la celda para poder pintarla con el metodo fill mas delante
                        valor_aplicacion1 = hoja_activa_busqueda_general_personas.cell(row=contador_busqueda_general_personas+contador_aplicaciones,column=4)

                        #aqui insertamos el usuario relacionado con esta aplicacion
                        hoja_activa_busqueda_general_personas.cell(row=contador_busqueda_general_personas+contador_aplicaciones,column=5).value = str(info_aplicacion.usuario)
                        #aqui vamos a guardar en una variable la celda para poder pintarla con el metodo fill mas delante
                        valor_aplicacion2 = hoja_activa_busqueda_general_personas.cell(row=contador_busqueda_general_personas+contador_aplicaciones,column=5)

                        #aqui insertamos el ticket de creacion del usuario de la aplicacion
                        hoja_activa_busqueda_general_personas.cell(row=contador_busqueda_general_personas+contador_aplicaciones,column=6).value = str(info_aplicacion.ticket)
                        #aqui vamos a guardar en una variable la celda para poder pintarla con el metodo fill mas delante
                        valor_aplicacion3 = hoja_activa_busqueda_general_personas.cell(row=contador_busqueda_general_personas+contador_aplicaciones,column=6)

                        #luego aqui evaluamos si la variable color es 0 entonces que lo pinte de un color diferente que si la variable color fuera 1
                        #y pues de esta manera se pintada cada fila de un color diferente
                        if color == 0:
                            valor_aplicacion1.fill = PatternFill("solid", fgColor="bdbdbd")
                            valor_aplicacion2.fill = PatternFill("solid", fgColor="bdbdbd")
                            valor_aplicacion3.fill = PatternFill("solid", fgColor="bdbdbd")
                        else:
                            valor_aplicacion1.fill = PatternFill("solid", fgColor="a3a3a3")
                            valor_aplicacion2.fill = PatternFill("solid", fgColor="a3a3a3")
                            valor_aplicacion3.fill = PatternFill("solid", fgColor="a3a3a3")


                        #y al contador le sumamos uno, el cual sera la fila, o sea que sera una fila mas abajo
                        contador_aplicaciones=contador_aplicaciones + 1
                #al final le sumamos al contadorgeneral el del principio, el numero de aplicaciones de la persona, para que en el momento de que se registre la nueva aplicacion, pues se registr abajo de la insertada anteriormente y no haya ni espacios ni problemas de ubicacion
                contador_busqueda_general_personas = contador_busqueda_general_personas + registro.aplicaciones.all().count()


                #luego aqui evaluamos si la variable color es 0 entonces que lo pinte de un color diferente que si la variable color fuera 1
                #y pues de esta manera se pintada cada fila de un color diferente
                if color == 0:
                    valor1.fill = PatternFill("solid", fgColor="bdbdbd")
                    valor2.fill = PatternFill("solid", fgColor="bdbdbd")
                    valor3.fill = PatternFill("solid", fgColor="bdbdbd")
                    color = 1
                else:
                    valor1.fill = PatternFill("solid", fgColor="a3a3a3")
                    valor2.fill = PatternFill("solid", fgColor="a3a3a3")
                    valor3.fill = PatternFill("solid", fgColor="a3a3a3")
                    color = 0


            #despues de registrar los datos en la tabla, vamos a ponerle nombre al archivo
            nombre_archivo_busqueda_general_personas = "Reporte_de_personas.xlsx"

            #despues vamos a definir con los siguientes comandos, el metodo de respuesta que se enviara
            #aqui le indicamos que la respuesta va a ser un archivo de microsoft excel
            response = HttpResponse(content_type="application/ms-excel")
            #aqui especificamos el nombre del archivo
            contenido = "attachment; filename={0}".format(nombre_archivo_busqueda_general_personas)
            #aqui le declaramos a la respuesta el nombre del archivo
            response["Content-Disposition"] = contenido
            #aqui guardamos el archivo en la respuesta
            libro_busqueda_general_personas.save(response)
            # y por utlimo devolvemos la respuesta que sera el archivo
            return response
        else:
            #LUEGO DE HABER HECHO LOS FILTROS, EVALUAREMOS SI EL REQUEST TRAE ALGUNA SOLICITUD DE ORDENADO POR ALGUN CAMPO
            return render(request, 'listar_personas/ajax_busquedas_con_filtros.html', {'lista':lista})




#esta es la funcion de eliminar personas, le pasamos por parametro la solicitud
def eliminar_persona(request):
    #si el metodo de entrada es GET, entonces que haga lo siguiente
    if request.method == 'GET':
        #aqui vamos a guardar en una variable la persona que queremos eliminar, segun el id que obtenemos enviado por medio de AJAX hasta aqui
        persona_a_eliminar = persona.objects.get(id=request.GET.get('id'))

        #aqui vamos  ahacer el diccionario con los datos para el historico de la elimincacion de la persona
        aplicaciones = ""
        #aqui iteramos las aplicaciones y las guardamos y las partimos por comas
        for apli in persona_a_eliminar.aplicaciones.all():
            aplicaciones+=str(apli)+str(",")
        #aqui guardamos en un diccionario todos los valores necesarios para el formulario de historico de eliminacion de persona
        values_persona_a_eliminar = {
            'persona' : persona_a_eliminar.nombre_completo,
            'aplicaciones' : aplicaciones,
            'cargo' : persona_a_eliminar.cargo,
            'ubicacion' : persona_a_eliminar.ubicacion,
        }
        #aqui vamos a guardar el nombre de la persona eliminada para mostrarlo en el template
        persona_eliminada = persona_a_eliminar.nombre_completo


        #aqui vamos a hacer el registro de la persona eliminada en el registro de historico de eliminacion de personas
        form_historico_eliminacion_persona = HistoricoForm_eliminacion_persona(values_persona_a_eliminar)
        #si el formulario es valido lo guarda, y despues elimina la persona
        if form_historico_eliminacion_persona.is_valid():
            form_historico_eliminacion_persona.save()
            #luego lo eliminamos
            persona_a_eliminar.delete()

        else:
            print("el formulario de historico de elimincacion de persona no es valido")

        #y aqui hacemos una variable de que si fue eliminado entonces que ponga que si, esto lo utilizaremos en el template
        borrado = "si"

        #aqui vamos a devolver un renderizado, con la solicitud, al template especificado en el segundo parametro y de tercer parametro le pasamos las variables que uqeremos manipulas en el template, en este caso va ha ser borrado
        return render(request, 'listar_personas/ajax_eliminar_personas.html', {'borrado': borrado, 'persona': persona_eliminada})



#esta es la funcion de editar persona, le pasamos por parametro la solicitud
def editar_persona(request):
    #si la peticion es GET, se obtiene el id de la perosna a editar por medio de el request GET
    if request.method == "GET":
        persona_a_editar = persona.objects.get(id=request.GET.get('id_persona_a_editar'))
    # si es metodo POST, se obtiene por el request POST
    elif request.method == "POST":
        persona_a_editar = persona.objects.get(id=request.POST.get('id'))

    #si la solicitud es POST, empezara a hacer la validacion del formulario y el guardado de los cambios indicados
    if request.method == "POST":
        #instanciamos el formualrio con los datos de la request.POST y la instancia de la persona va a ser la persona a editar
        form = EditarPersonaForm(request.POST, instance=persona_a_editar)
        #si el formulario es valido
        if form.is_valid():
            #entonces que lo guarde
            persona_editada = form.save()
            print("guardando persona")
            form = persona_editada
        else:
            print("el formulario no es valido")
        print("hace el guardado de la edicion de la persona")
        # y aqui le indicamos una variable con el tipo de metodo y esta variable la vamos a utilizar en el template
        metodo = "POST"
    # de lo contrario a que sea un metodo POST, entonces hata el formulario con la instancia de la persona a editar y la variable metodo sera GET, y recuerden que esta la manipularemos en el template
    else:
        print(persona_a_editar)
        #instanciamos el formulario con la persona a editar
        form = EditarPersonaForm(instance=persona_a_editar)
        # la variavle metodo le asignamos el valor de GET
        metodo = "GET"
    #aqui retornamos un renderizado, le indicamos el HTML en el cual vamos a poner la informacion y le pasamos en un diccionario el formulario y la variable metodo
    return render(request, 'listar_personas/ajax_editar_persona.html', {'form': form, 'metodo': metodo})


#aqui hacemos la funcion que manejara los datos y enviara los mismos de eliminar aplicaciones en personas, recibira como parametro los datos de la solicitud
def eliminar_aplicacion_persona(request):
    #esta es la variable, que cambiara de valor si ingrrsamos en metodo POST o GET, y este valor, lo vamos a usar en el template
    metodo = "";
    #aqui lo que hacemos es una condicion que si el metodo de la solicitud es GET pues que obtenga el id de esa manera y de lo contrario que lo obtenga por medio del metodo POST
    if request.method == "GET":
        id_persona_a_eliminar_aplicacion = persona.objects.get(id=request.GET.get("id_persona_a_eliminar_aplicacion"))
    else:
        id_persona_a_eliminar_aplicacion = persona.objects.get(id=request.POST.get("id_persona_a_eliminar_aplicacion"))

    #aqui hacemos otra condicional, la cual dice que si es metodo POST, que haga la validacion del formulario y guarde los cambios de eliminacion y el del historico
    if request.method == "POST":
        #la variable metodo sera POST
        metodo = "POST"
        print(request.POST)
        print("hace la valiacion de datos por el metodo POST y si estan bien guarda los cambios")
        #primero vamos a traer el usuario de la aplicacion que queremos eliminar, por que es un campo requerido para el historico de eliminacion de la aplicacion
        usuario_aplicacion_a_eliminar = usuario.objects.get(persona_relacionada=request.POST.get('id_persona_a_eliminar_aplicacion'), aplicacion_relacionada=request.POST.get('aplicacion_a_eliminar'))

        #vamos a hacer el diccionario con los datos que requerimos para el formulario de historico de eliminacion de aplicacion.
        dic_historico_eliminacion_aplicacion = {}
        #empezamos metiendo en orden los campos requeridos
        dic_historico_eliminacion_aplicacion['csrfmiddlewaretoken'] = request.POST.get('csrfmiddlewaretoken')
        dic_historico_eliminacion_aplicacion['ticket'] = [request.POST.get('ticket_eliminacion_aplicacion_persona')]
        dic_historico_eliminacion_aplicacion['persona_relacionada'] = [id_persona_a_eliminar_aplicacion.nombre_completo]
        dic_historico_eliminacion_aplicacion['aplicacion_relacionada'] = [id_persona_a_eliminar_aplicacion.aplicaciones.get(id=request.POST.get('aplicacion_a_eliminar'))]
        dic_historico_eliminacion_aplicacion['usuario'] = [usuario_aplicacion_a_eliminar]
        print(dic_historico_eliminacion_aplicacion)

        #ahora vamos a hacer que el diccionario sea un diccionario de consulta, para que sea aceptado por el formulario de AplicacionForm_eliminacion
        data_form_historico_eliminacion_aplicacion = QueryDict('',mutable=True)
        data_form_historico_eliminacion_aplicacion.update(MultiValueDict(dic_historico_eliminacion_aplicacion))

        #ahora vamos a hacer la validacion del formulario y si esta bien vamos a guardarlo
        form_historico_eliminacion_aplicacion = AplicacionForm_eliminacion(data_form_historico_eliminacion_aplicacion)
        if form_historico_eliminacion_aplicacion.is_valid():
            form_historico_eliminacion_aplicacion.save()
            print("guardando historico de eliminacion aplicacion")

            #ahora vamos a hacer la eliminacion del usuario en la persona
            usuario_aplicacion_a_eliminar.delete()
            print("eliminacion del usuario")

            #ahora aqui vamos a hacer la eliminacion de la aplicacion en el modelo que este relacionado con la persona con el modelo aplicaciones, ya que lo que hicimos fue eliminar la relacion que habia entre la persona con el usaurio
            #ahora falta eliminar la relacion que hay entre la persona con la aplicacion
            #asi que lo que vamos a hacer es lo mismo que hicimos con el usuario
            #aqui guardamos en una variable la persona a la que queremos eliminarle la aplicacion
            aplicacion_a_eliminar_persona = persona.objects.get(id=request.POST.get('id_persona_a_eliminar_aplicacion'))
            #luego aqui vamos a buscar la aplicacion relacionada con la persona por medio del Id y vamos a remover la relacion
            aplicacion_a_eliminar_persona.aplicaciones.remove(aplicacion.objects.get(id=request.POST.get('aplicacion_a_eliminar')))

            #aqui le asignamos el valor a la variable metodo de POST, para asi manipularlo en la template
            metodo = "POST"

            #aqui le asigmnamos a la variable de form_eliminacion_aplicacion el valor de la persona que le eliminamos la aplicacion, para que en el renderizado, pues tenga los valores nuevos de la persona
            form_eliminacion_aplicacion = persona.objects.get(id=request.POST.get('id_persona_a_eliminar_aplicacion'))
        else:
            print("el formulario no es valido")

    else:
        #de lo contrario, vamos a crear el formulario de eliminacion aplicaciones y lo llenamos con la informacion de la persona a modificar
        #este codigo es un poco confuso, ya que estamos utilzando un formulario para agregar aplicacion, pero este tambien nos puede servir para eliminar las aplicaciones
        form_eliminacion_aplicacion = PersonaForm_addaplicacion(instance=id_persona_a_eliminar_aplicacion)

        #aqui vamos a modificar la variable de si es POST o GET, y como estamos dentro del metodo GET, la variable, sera de string GET
        metodo = "GET"


    #aqui vamos a retornar un renderizado que llevara 3 parametros, el primer son los datos de la solicitud, que es la request
    #el segundo parametro es el template al cual vamos a enviar la informacion
    #y el tercer parametro son los datos encapsulados en un diccionarios, los cuales vamos a utilizar en el template
    return render(request, 'listar_personas/ajax_eliminar_aplicaciones_persona.html', {'form': form_eliminacion_aplicacion, 'metodo': metodo})


#AQUI HAREMOS LA VISTA DE LISTAR LAS EL HISTORICO DE PERSONAS CREADAS
def listar_historico_personas_creadas(request):
    historico_personas_creadas = historico_creacion_persona.objects.all()
    return render(request, 'historicos/ajax_historico_personas_creadas.html', {'personas':historico_personas_creadas})
#AQUI HAREMOS LA VISTA DE LISTAR LAS EL HISTORICO DE PERSONAS CREADAS


#AQUI VAMOS A HACER LA VISTA DE EL REPORTE HISTORICO DE PERSONAS CREADAS
def busqueda_historico_personas_creadas(request):
    #aqui decimos si el metodo es GET, que comienze la funcion
    if request.method == "GET":
        #aqui hacemos la obtencion de todos los elementos de el modelo historico y los metemos es una variable
        historico_personas_creadas = historico_creacion_persona.objects.all()

        #luego aqui hacemos la condicional de que si existen las variables fecha inicio y fecha fin, entonces que filtre la variable que contiene los registros de historicos de creacion de personas por las fechas indicadas
        if request.GET.get('fecha_inicio') and request.GET.get('fecha_fin'):
            #aqui guardamos en una variable la fecha de inicio obtenida desde AJAX
            fecha_inicio = request.GET.get('fecha_inicio')
            #aqui guardamos en una variable la fecha fin
            fecha_fin = request.GET.get('fecha_fin')
            #aqui hacemos el filtrado de los registros por medio de las fechas indicadas
            historico_personas_creadas = historico_personas_creadas.filter(hora__range=(fecha_inicio,fecha_fin))

        #luego aqui hacemos otra condicional, dira que si el valor en la url query_general existe, quiere decir que se mando un valor de busqueda general, entonces tendremos que hacerlo
        if request.GET.get('query_general'):

            #aqui declaramos en una variable el valor de query general.
            consulta_general_historico_personas_creadas = request.GET.get('query_general')
            #aqui hacemos una variable que contendra unos campos de busqueda con la funcion Q, en la cual buscaremos en cada campo el valor indicado en la variable query_general
            qset_historico_personas_creadas = (
                Q(hora__icontains=consulta_general_historico_personas_creadas) |
                Q(persona__icontains=consulta_general_historico_personas_creadas) |
                Q(aplicaciones__icontains=consulta_general_historico_personas_creadas) |
                Q(cargo__icontains=consulta_general_historico_personas_creadas) |
                Q(ubicacion__icontains=consulta_general_historico_personas_creadas)
            )
            #luego aqui hace el filtrado de los registro por medio de la consulta y el ultimo parametro es que si encuentra algun registro duplicado que no lo muestre
            historico_personas_creadas = historico_personas_creadas.filter(qset_historico_personas_creadas).distinct()

        #aqui vamos a hacer el codigo para poder validar si se quiere exportar a excel o no
        #validamos si el request tiene la variable exportacion_a_excel y si es asi entonces cogemos la variable de hemos estado modificando con los filtros de busqeuda generaly de fechas y la utilizaremos para poder ingresarlo en el excel
        if request.GET.get('exportacion_a_excel'):
            #entonces aqui vamos a comenzar a crear el libro con los siguiente comandos
            #instanciamos el libro y li llamamos libro_historico_personas_creadas
            libro_historico_personas_creadas = Workbook()
            #luego nos toca definir una hoja como activa para comenzar a editarla, entonces por defecto vamos a activar la hoja 1 por defecto
            hoja_activa_historico_personas_creadas = libro_historico_personas_creadas.active

            #luego con la hoja activa vamos a trabajar
            #primero vamos a poner el nombre de la tabla con el siguiente parametro
            hoja_activa_historico_personas_creadas['A1'] = "Reporte historico personas creadas"
            #ahora vamos a juntar algunas celdas para que se vea que el titulo abarca toda la tabla
            hoja_activa_historico_personas_creadas.merge_cells('A1:E1')

            #ahora vamos a crear los campos de la tabla, o sea los titulos de cada columna
            hoja_activa_historico_personas_creadas['A2'] = "Hora creacion"
            hoja_activa_historico_personas_creadas['B2'] = "Persona"
            hoja_activa_historico_personas_creadas['C2'] = "Aplicaciones"
            hoja_activa_historico_personas_creadas['D2'] = "Cargo"
            hoja_activa_historico_personas_creadas['E2'] = "Ubicacion"

            #Ahora vamos a crear un contador para recorrer todos los registros de la variable que tiene los registros de la tabla de historicos de personas creadas
            contador_historico_personas_creadas=3

            #ahora vamos a empezar a insertar todos los registros en la hoja segun el contador, el contador va a ser la fila actual y se incrementara cada vez que se registre una nueva persona
            for registro in historico_personas_creadas:
                hoja_activa_historico_personas_creadas.cell(row=contador_historico_personas_creadas,column=1).value = registro.hora
                hoja_activa_historico_personas_creadas.cell(row=contador_historico_personas_creadas,column=2).value = registro.persona
                hoja_activa_historico_personas_creadas.cell(row=contador_historico_personas_creadas,column=3).value = registro.aplicaciones
                hoja_activa_historico_personas_creadas.cell(row=contador_historico_personas_creadas,column=4).value = registro.cargo
                hoja_activa_historico_personas_creadas.cell(row=contador_historico_personas_creadas,column=5).value = registro.ubicacion
                contador_historico_personas_creadas = contador_historico_personas_creadas + 1

            #despues de registrar los datos en la tabla, vamos a ponerle nombre al archivo
            nombre_archivo_historico_personas_creadas = "HistoricoPersonasCreadas.xlsx"

            #despues vamos a definir con los siguientes comandos, el metodo de respuesta que se enviara
            #aqui le indicamos que la respuesta va a ser un archivo de microsoft excel
            response = HttpResponse(content_type="application/ms-excel")
            #aqui especificamos el nombre del archivo
            contenido = "attachment; filename={0}".format(nombre_archivo_historico_personas_creadas)
            #aqui le declaramos a la respuesta el nombre del archivo
            response["Content-Disposition"] = contenido
            #aqui guardamos el archivo en la respuesta
            libro_historico_personas_creadas.save(response)
            # y por utlimo devolvemos la respuesta que sera el archivo
            return response



    #y aqui por ultimo hace el renderizado de la informacion al template y le pasamos los parametros y las variables hechas en el codigo anterior
    return render(request, 'historicos/ajax_busquedas_historicos_personas.html', {'personas': historico_personas_creadas})











#AQUI HAREMOS LA VISTA DE LISTAR LAS EL HISTORICO DE PERSONAS ELIMINADAS
def listar_historico_personas_eliminadas(request):
    historico_personas_eliminadas = historico_eliminacion_persona.objects.all()
    return render(request, 'historicos/ajax_historico_personas_eliminadas.html', {'personas':historico_personas_eliminadas})
#AQUI HAREMOS LA VISTA DE LISTAR LAS EL HISTORICO DE PERSONAS ELIMINADAS



#AQUI VAMOS A HACER LA VISTA DE EL REPORTE HISTORICO DE PERSONAS CREADAS
def busqueda_historico_personas_eliminadas(request):
    #aqui decimos si el metodo es GET, que comienze la funcion
    if request.method == "GET":
        #aqui hacemos la obtencion de todos los elementos de el modelo historico y los metemos es una variable
        historico_personas_eliminadas = historico_eliminacion_persona.objects.all()

        #luego aqui hacemos la condicional de que si existen las variables fecha inicio y fecha fin, entonces que filtre la variable que contiene los registros de historicos de creacion de personas por las fechas indicadas
        if request.GET.get('fecha_inicio') and request.GET.get('fecha_fin'):
            #aqui guardamos en una variable la fecha de inicio obtenida desde AJAX
            fecha_inicio = request.GET.get('fecha_inicio')
            #aqui guardamos en una variable la fecha fin
            fecha_fin = request.GET.get('fecha_fin')
            #aqui hacemos el filtrado de los registros por medio de las fechas indicadas
            historico_personas_eliminadas = historico_personas_eliminadas.filter(hora__range=(fecha_inicio,fecha_fin))

        #luego aqui hacemos otra condicional, dira que si el valor en la url query_general existe, quiere decir que se mando un valor de busqueda general, entonces tendremos que hacerlo
        if request.GET.get('query_general'):

            #aqui declaramos en una variable el valor de query general.
            consulta_general_historico_personas_eliminadas = request.GET.get('query_general')
            #aqui hacemos una variable que contendra unos campos de busqueda con la funcion Q, en la cual buscaremos en cada campo el valor indicado en la variable query_general
            qset_historico_personas_eliminadas = (
                Q(hora__icontains=consulta_general_historico_personas_eliminadas) |
                Q(persona__icontains=consulta_general_historico_personas_eliminadas) |
                Q(aplicaciones__icontains=consulta_general_historico_personas_eliminadas) |
                Q(cargo__icontains=consulta_general_historico_personas_eliminadas) |
                Q(ubicacion__icontains=consulta_general_historico_personas_eliminadas)
            )
            #luego aqui hace el filtrado de los registro por medio de la consulta y el ultimo parametro es que si encuentra algun registro duplicado que no lo muestre
            historico_personas_eliminadas = historico_personas_eliminadas.filter(qset_historico_personas_eliminadas).distinct()

        #aqui vamos a hacer el codigo para poder validar si se quiere exportar a excel o no
        #validamos si el request tiene la variable exportacion_a_excel y si es asi entonces cogemos la variable de hemos estado modificando con los filtros de busqeuda generaly de fechas y la utilizaremos para poder ingresarlo en el excel
        if request.GET.get('exportacion_a_excel'):
            #entonces aqui vamos a comenzar a crear el libro con los siguiente comandos
            #instanciamos el libro y li llamamos libro_historico_personas_creadas
            libro_historico_personas_eliminadas = Workbook()
            #luego nos toca definir una hoja como activa para comenzar a editarla, entonces por defecto vamos a activar la hoja 1 por defecto
            hoja_activa_historico_personas_eliminadas = libro_historico_personas_eliminadas.active

            #luego con la hoja activa vamos a trabajar
            #primero vamos a poner el nombre de la tabla con el siguiente parametro
            hoja_activa_historico_personas_eliminadas['A1'] = "Reporte historico personas eliminadas"
            #ahora vamos a juntar algunas celdas para que se vea que el titulo abarca toda la tabla
            hoja_activa_historico_personas_eliminadas.merge_cells('A1:E1')

            #ahora vamos a crear los campos de la tabla, o sea los titulos de cada columna
            hoja_activa_historico_personas_eliminadas['A2'] = "Hora creacion"
            hoja_activa_historico_personas_eliminadas['B2'] = "Persona"
            hoja_activa_historico_personas_eliminadas['C2'] = "Aplicaciones"
            hoja_activa_historico_personas_eliminadas['D2'] = "Cargo"
            hoja_activa_historico_personas_eliminadas['E2'] = "Ubicacion"

            #Ahora vamos a crear un contador para recorrer todos los registros de la variable que tiene los registros de la tabla de historicos de personas creadas
            contador_historico_personas_eliminadas=3

            #ahora vamos a empezar a insertar todos los registros en la hoja segun el contador, el contador va a ser la fila actual y se incrementara cada vez que se registre una nueva persona
            for registro in historico_personas_eliminadas:
                hoja_activa_historico_personas_eliminadas.cell(row=contador_historico_personas_eliminadas,column=1).value = registro.hora
                hoja_activa_historico_personas_eliminadas.cell(row=contador_historico_personas_eliminadas,column=2).value = registro.persona
                hoja_activa_historico_personas_eliminadas.cell(row=contador_historico_personas_eliminadas,column=3).value = registro.aplicaciones
                hoja_activa_historico_personas_eliminadas.cell(row=contador_historico_personas_eliminadas,column=4).value = registro.cargo
                hoja_activa_historico_personas_eliminadas.cell(row=contador_historico_personas_eliminadas,column=5).value = registro.ubicacion
                contador_historico_personas_eliminadas = contador_historico_personas_eliminadas + 1

            #despues de registrar los datos en la tabla, vamos a ponerle nombre al archivo
            nombre_archivo_historico_personas_eliminadas = "HistoricoPersonasEliminadas.xlsx"

            #despues vamos a definir con los siguientes comandos, el metodo de respuesta que se enviara
            #aqui le indicamos que la respuesta va a ser un archivo de microsoft excel
            response = HttpResponse(content_type="application/ms-excel")
            #aqui especificamos el nombre del archivo
            contenido = "attachment; filename={0}".format(nombre_archivo_historico_personas_eliminadas)
            #aqui le declaramos a la respuesta el nombre del archivo
            response["Content-Disposition"] = contenido
            #aqui guardamos el archivo en la respuesta
            libro_historico_personas_eliminadas.save(response)
            # y por utlimo devolvemos la respuesta que sera el archivo
            return response



    #y aqui por ultimo hace el renderizado de la informacion al template y le pasamos los parametros y las variables hechas en el codigo anterior
    return render(request, 'historicos/ajax_busquedas_historicos_personas.html', {'personas': historico_personas_eliminadas})











#AQUI HAREMOS LA VISTA DE LISTAR LAS EL HISTORICO DE APLICACIONES CREADAS
def listar_historico_aplicaciones_creadas(request):
    historico_aplicaciones_creadas = creacion_aplicacion.objects.all()
    return render(request, 'historicos/ajax_historico_aplicaciones_creadas.html', {'personas':historico_aplicaciones_creadas})
#AQUI HAREMOS LA VISTA DE LISTAR LAS EL HISTORICO DE APLICACIONES CREADAS



#AQUI VAMOS A HACER LA VISTA DE EL REPORTE HISTORICO DE PERSONAS CREADAS
def busqueda_historico_aplicaciones_creadas(request):
    #aqui decimos si el metodo es GET, que comienze la funcion
    if request.method == "GET":
        #aqui hacemos la obtencion de todos los elementos de el modelo historico y los metemos es una variable
        historico_aplicaciones_creadas = creacion_aplicacion.objects.all()

        #luego aqui hacemos la condicional de que si existen las variables fecha inicio y fecha fin, entonces que filtre la variable que contiene los registros de historicos de creacion de personas por las fechas indicadas
        if request.GET.get('fecha_inicio') and request.GET.get('fecha_fin'):
            #aqui guardamos en una variable la fecha de inicio obtenida desde AJAX
            fecha_inicio = request.GET.get('fecha_inicio')
            #aqui guardamos en una variable la fecha fin
            fecha_fin = request.GET.get('fecha_fin')
            #aqui hacemos el filtrado de los registros por medio de las fechas indicadas
            historico_aplicaciones_creadas = historico_aplicaciones_creadas.filter(hora__range=(fecha_inicio,fecha_fin))

        #luego aqui hacemos otra condicional, dira que si el valor en la url query_general existe, quiere decir que se mando un valor de busqueda general, entonces tendremos que hacerlo
        if request.GET.get('query_general'):

            #aqui declaramos en una variable el valor de query general.
            consulta_general_historico_aplicaciones_creadas = request.GET.get('query_general')
            #aqui hacemos una variable que contendra unos campos de busqueda con la funcion Q, en la cual buscaremos en cada campo el valor indicado en la variable query_general
            qset_historico_aplicaciones_creadas = (
                Q(ticket__icontains=consulta_general_historico_aplicaciones_creadas) |
                Q(persona_relacionada__icontains=consulta_general_historico_aplicaciones_creadas) |
                Q(aplicacion_relacionada__icontains=consulta_general_historico_aplicaciones_creadas) |
                Q(usuario__icontains=consulta_general_historico_aplicaciones_creadas) |
                Q(hora__icontains=consulta_general_historico_aplicaciones_creadas)
            )
            #luego aqui hace el filtrado de los registro por medio de la consulta y el ultimo parametro es que si encuentra algun registro duplicado que no lo muestre
            historico_aplicaciones_creadas = historico_aplicaciones_creadas.filter(qset_historico_aplicaciones_creadas).distinct()

        #aqui vamos a hacer el codigo para poder validar si se quiere exportar a excel o no
        #validamos si el request tiene la variable exportacion_a_excel y si es asi entonces cogemos la variable de hemos estado modificando con los filtros de busqeuda generaly de fechas y la utilizaremos para poder ingresarlo en el excel
        if request.GET.get('exportacion_a_excel'):
            #entonces aqui vamos a comenzar a crear el libro con los siguiente comandos
            #instanciamos el libro y li llamamos libro_historico_personas_creadas
            libro_historico_aplicaciones_creadas = Workbook()
            #luego nos toca definir una hoja como activa para comenzar a editarla, entonces por defecto vamos a activar la hoja 1 por defecto
            hoja_activa_historico_aplicaciones_creadas = libro_historico_aplicaciones_creadas.active

            #luego con la hoja activa vamos a trabajar
            #primero vamos a poner el nombre de la tabla con el siguiente parametro
            hoja_activa_historico_aplicaciones_creadas['A1'] = "Reporte historico aplicaciones creadas"
            #ahora vamos a juntar algunas celdas para que se vea que el titulo abarca toda la tabla
            hoja_activa_historico_aplicaciones_creadas.merge_cells('A1:E1')

            #ahora vamos a crear los campos de la tabla, o sea los titulos de cada columna
            hoja_activa_historico_aplicaciones_creadas['A2'] = "Hora creacion"
            hoja_activa_historico_aplicaciones_creadas['B2'] = "Persona"
            hoja_activa_historico_aplicaciones_creadas['C2'] = "Aplicacion"
            hoja_activa_historico_aplicaciones_creadas['D2'] = "Usuario"
            hoja_activa_historico_aplicaciones_creadas['E2'] = "Ticket creacion"

            #Ahora vamos a crear un contador para recorrer todos los registros de la variable que tiene los registros de la tabla de historicos de personas creadas
            contador_historico_aplicaciones_creadas=3

            #ahora vamos a empezar a insertar todos los registros en la hoja segun el contador, el contador va a ser la fila actual y se incrementara cada vez que se registre una nueva persona
            for registro in historico_aplicaciones_creadas:
                hoja_activa_historico_aplicaciones_creadas.cell(row=contador_historico_aplicaciones_creadas,column=1).value = registro.hora
                hoja_activa_historico_aplicaciones_creadas.cell(row=contador_historico_aplicaciones_creadas,column=2).value = registro.persona_relacionada
                hoja_activa_historico_aplicaciones_creadas.cell(row=contador_historico_aplicaciones_creadas,column=3).value = registro.aplicacion_relacionada
                hoja_activa_historico_aplicaciones_creadas.cell(row=contador_historico_aplicaciones_creadas,column=4).value = registro.usuario
                hoja_activa_historico_aplicaciones_creadas.cell(row=contador_historico_aplicaciones_creadas,column=5).value = registro.ticket
                contador_historico_aplicaciones_creadas = contador_historico_aplicaciones_creadas + 1

            #despues de registrar los datos en la tabla, vamos a ponerle nombre al archivo
            nombre_archivo_historico_aplicaciones_creadas = "HistoricoAplicacionesCreadas.xlsx"

            #despues vamos a definir con los siguientes comandos, el metodo de respuesta que se enviara
            #aqui le indicamos que la respuesta va a ser un archivo de microsoft excel
            response = HttpResponse(content_type="application/ms-excel")
            #aqui especificamos el nombre del archivo
            contenido = "attachment; filename={0}".format(nombre_archivo_historico_aplicaciones_creadas)
            #aqui le declaramos a la respuesta el nombre del archivo
            response["Content-Disposition"] = contenido
            #aqui guardamos el archivo en la respuesta
            libro_historico_aplicaciones_creadas.save(response)
            # y por utlimo devolvemos la respuesta que sera el archivo
            return response



    #y aqui por ultimo hace el renderizado de la informacion al template y le pasamos los parametros y las variables hechas en el codigo anterior
    return render(request, 'historicos/ajax_busquedas_historicos_aplicaciones.html', {'personas': historico_aplicaciones_creadas})









#AQUI HAREMOS LA VISTA DE LISTAR LAS EL HISTORICO DE APLICACIONES ELIMINADAS
def listar_historico_aplicaciones_eliminadas(request):
    historico_aplicaciones_eliminadas = eliminacion_aplicacion.objects.all()
    return render(request, 'historicos/ajax_historico_aplicaciones_eliminadas.html', {'personas':historico_aplicaciones_eliminadas})
#AQUI HAREMOS LA VISTA DE LISTAR LAS EL HISTORICO DE APLICACIONES ELIMINADAS



#AQUI VAMOS A HACER LA VISTA DE EL REPORTE HISTORICO DE PERSONAS ELIMINADAS
def busqueda_historico_aplicaciones_eliminadas(request):
    #aqui decimos si el metodo es GET, que comienze la funcion
    if request.method == "GET":
        #aqui hacemos la obtencion de todos los elementos de el modelo historico y los metemos es una variable
        historico_aplicaciones_eliminadas = eliminacion_aplicacion.objects.all()

        #luego aqui hacemos la condicional de que si existen las variables fecha inicio y fecha fin, entonces que filtre la variable que contiene los registros de historicos de creacion de personas por las fechas indicadas
        if request.GET.get('fecha_inicio') and request.GET.get('fecha_fin'):
            #aqui guardamos en una variable la fecha de inicio obtenida desde AJAX
            fecha_inicio = request.GET.get('fecha_inicio')
            #aqui guardamos en una variable la fecha fin
            fecha_fin = request.GET.get('fecha_fin')
            #aqui hacemos el filtrado de los registros por medio de las fechas indicadas
            historico_aplicaciones_eliminadas = historico_aplicaciones_eliminadas.filter(hora__range=(fecha_inicio,fecha_fin))

        #luego aqui hacemos otra condicional, dira que si el valor en la url query_general existe, quiere decir que se mando un valor de busqueda general, entonces tendremos que hacerlo
        if request.GET.get('query_general'):

            #aqui declaramos en una variable el valor de query general.
            consulta_general_historico_aplicaciones_eliminadas = request.GET.get('query_general')
            #aqui hacemos una variable que contendra unos campos de busqueda con la funcion Q, en la cual buscaremos en cada campo el valor indicado en la variable query_general
            qset_historico_aplicaciones_eliminadas = (
                Q(ticket__icontains=consulta_general_historico_aplicaciones_eliminadas) |
                Q(persona_relacionada__icontains=consulta_general_historico_aplicaciones_eliminadas) |
                Q(aplicacion_relacionada__icontains=consulta_general_historico_aplicaciones_eliminadas) |
                Q(usuario__icontains=consulta_general_historico_aplicaciones_eliminadas) |
                Q(hora__icontains=consulta_general_historico_aplicaciones_eliminadas)
            )
            #luego aqui hace el filtrado de los registro por medio de la consulta y el ultimo parametro es que si encuentra algun registro duplicado que no lo muestre
            historico_aplicaciones_eliminadas = historico_aplicaciones_eliminadas.filter(qset_historico_aplicaciones_eliminadas).distinct()

        #aqui vamos a hacer el codigo para poder validar si se quiere exportar a excel o no
        #validamos si el request tiene la variable exportacion_a_excel y si es asi entonces cogemos la variable de hemos estado modificando con los filtros de busqeuda generaly de fechas y la utilizaremos para poder ingresarlo en el excel
        if request.GET.get('exportacion_a_excel'):
            #entonces aqui vamos a comenzar a crear el libro con los siguiente comandos
            #instanciamos el libro y li llamamos libro_historico_personas_creadas
            libro_historico_aplicaciones_eliminadas = Workbook()
            #luego nos toca definir una hoja como activa para comenzar a editarla, entonces por defecto vamos a activar la hoja 1 por defecto
            hoja_activa_historico_aplicaciones_eliminadas = libro_historico_aplicaciones_eliminadas.active

            #luego con la hoja activa vamos a trabajar
            #primero vamos a poner el nombre de la tabla con el siguiente parametro
            hoja_activa_historico_aplicaciones_eliminadas['A1'] = "Reporte historico aplicaciones eliminadas"
            #ahora vamos a juntar algunas celdas para que se vea que el titulo abarca toda la tabla
            hoja_activa_historico_aplicaciones_eliminadas.merge_cells('A1:E1')

            #ahora vamos a crear los campos de la tabla, o sea los titulos de cada columna
            hoja_activa_historico_aplicaciones_eliminadas['A2'] = "Hora eliminacion"
            hoja_activa_historico_aplicaciones_eliminadas['B2'] = "Persona"
            hoja_activa_historico_aplicaciones_eliminadas['C2'] = "Aplicacion"
            hoja_activa_historico_aplicaciones_eliminadas['D2'] = "Usuario"
            hoja_activa_historico_aplicaciones_eliminadas['E2'] = "Ticket eliminacion"

            #Ahora vamos a crear un contador para recorrer todos los registros de la variable que tiene los registros de la tabla de historicos de personas creadas
            contador_historico_aplicaciones_eliminadas=3

            #ahora vamos a empezar a insertar todos los registros en la hoja segun el contador, el contador va a ser la fila actual y se incrementara cada vez que se registre una nueva persona
            for registro in historico_aplicaciones_eliminadas:
                hoja_activa_historico_aplicaciones_eliminadas.cell(row=contador_historico_aplicaciones_eliminadas,column=1).value = registro.hora
                hoja_activa_historico_aplicaciones_eliminadas.cell(row=contador_historico_aplicaciones_eliminadas,column=2).value = registro.persona_relacionada
                hoja_activa_historico_aplicaciones_eliminadas.cell(row=contador_historico_aplicaciones_eliminadas,column=3).value = registro.aplicacion_relacionada
                hoja_activa_historico_aplicaciones_eliminadas.cell(row=contador_historico_aplicaciones_eliminadas,column=4).value = registro.usuario
                hoja_activa_historico_aplicaciones_eliminadas.cell(row=contador_historico_aplicaciones_eliminadas,column=5).value = registro.ticket
                contador_historico_aplicaciones_eliminadas = contador_historico_aplicaciones_eliminadas + 1

            #despues de registrar los datos en la tabla, vamos a ponerle nombre al archivo
            nombre_archivo_historico_aplicaciones_eliminadas = "HistoricoAplicacionesEliminadas.xlsx"

            #despues vamos a definir con los siguientes comandos, el metodo de respuesta que se enviara
            #aqui le indicamos que la respuesta va a ser un archivo de microsoft excel
            response = HttpResponse(content_type="application/ms-excel")
            #aqui especificamos el nombre del archivo
            contenido = "attachment; filename={0}".format(nombre_archivo_historico_aplicaciones_eliminadas)
            #aqui le declaramos a la respuesta el nombre del archivo
            response["Content-Disposition"] = contenido
            #aqui guardamos el archivo en la respuesta
            libro_historico_aplicaciones_eliminadas.save(response)
            # y por utlimo devolvemos la respuesta que sera el archivo
            return response



    #y aqui por ultimo hace el renderizado de la informacion al template y le pasamos los parametros y las variables hechas en el codigo anterior
    return render(request, 'historicos/ajax_busquedas_historicos_aplicaciones.html', {'personas': historico_aplicaciones_eliminadas})
